import pdfplumber
import pandas as pd
import re
from collections import defaultdict

# === Atur input/output
pdf_path = "LOGNUS5_VOY1-5.pdf"
output_csv = "Manifest_Lognus5.csv"
output_excel = "Manifest_Lognus5.xlsx"

# === Regex metadata
kapal_regex = re.compile(r"NAMA KAPAL\s*:\s*(.*?)\s*\(", re.IGNORECASE)
voyage_regex = re.compile(r"RUTE/VOYAGE\s*:\s*(\d+)", re.IGNORECASE)
pelabuhan_regex = re.compile(r"PELABUHAN MUAT\s*:\s*(.*?)\s+PEL.*BONGKAR\s*:\s*(.*?)\s", re.IGNORECASE)

def fix_multiline_cell(text):
    if not text:
        return ""
    lines = str(text).split("\n")
    result = ""
    for i, line in enumerate(lines):
        line = line.strip()
        if line.endswith("-"):
            result += line[:-1]  # hilangkan '-' dan sambung langsung
        else:
            result += line + " " if i < len(lines) - 1 else line
    return result.strip()

# === Fungsi bantu
def normalize_shipper(name):
    if not name:
        return ""
    name = re.sub(r"\bP\s*\.?\s*T\.?\b", "PT", name, flags=re.IGNORECASE)
    name = re.sub(r"\bC\s*\.?\s*V\.?\b", "CV", name, flags=re.IGNORECASE)
    name = re.sub(r"\b(PT|CV)\b(?!\.)", r"\1.", name)  # Tambah titik hanya jika belum ada
    return name.strip().upper()

# === Konversi format uang menjadi numerik
def parse_money(text):
    if not text:
        return 0
    try:
        text = re.sub(r"\.00$", "", text.strip())
        clean_number = re.sub(r"[^\d]", "", text)
        return int(clean_number)
    except:
        return 0

# === Proses utama
grouped_data = defaultdict(lambda: {
    "Quantity": 0,
    "Container": [],
    "Contents": [],
    "Measurement": [],
    "Weight": [],
    "Uang Tambang": 0
})

# Untuk menyimpan uang tambang berdasarkan (prefix, remark)
uang_tambang_per_grup = defaultdict(int)

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue

        kapal = kapal_regex.search(text).group(1).strip().upper() if kapal_regex.search(text) else ""
        voyage = voyage_regex.search(text).group(1).strip() if voyage_regex.search(text) else ""
        pel_bongkar = pelabuhan_regex.search(text).group(2).strip().upper() if pelabuhan_regex.search(text) else ""

        table = page.extract_table()
        if not table:
            continue

        last_values = [""] * 11
        for row in table:
            if not any(row): continue
            if len(row) < 2: continue

            # Skip baris header & field template
            header_keywords = ["no. of b/l", "no.konosemen", "shipper", "pengirim", "consignee", "penerima", "(1)", "(2)", "(3)"]

            # Skip baris header tabel
            if any(any(hdr in str(cell).lower() for hdr in header_keywords) for cell in row[:3]):
                continue

            # Juga abaikan baris jumlah
            if row[1] and "JUMLAH" in row[1].upper():
                continue

            # Wariskan nilai kosong dari baris atas
            row_extended = row + [""] * (11 - len(row))
            row = [row_extended[i] if row_extended[i] else last_values[i] for i in range(11)]
            last_values = row.copy()

            no_bl = row[0]

            raw_shipper = fix_multiline_cell(row[1])
            shipper = normalize_shipper(raw_shipper)

            raw_consignee = fix_multiline_cell(row[2])
            consignee = raw_consignee.strip().upper()

            quantity_raw, container, package, contents, meas, weight, prepaid_raw, remarks = row[3:11]

            containers = [c.strip() for c in str(container).split(",") if c.strip()]
            contents_list = [c.strip() for c in str(contents).split(",")] if contents else []
            meas_list = [m.strip() for m in str(meas).split(",")] if meas else []
            weight_list = [w.strip() for w in str(weight).split(",")] if weight else []

            uang_tambang = parse_money(prepaid_raw)

            remark_list = [r.strip() for r in remarks.split(",")] if remarks and "," in remarks else [remarks.strip() if remarks else ""]

            # Mapping kontainer berdasarkan keterangan
            remark_index = {}
            if len(remark_list) == len(containers):
                for i, r in enumerate(remark_list):
                    remark_index.setdefault(r, []).append(i)
            else:
                remark_index[remark_list[0]] = list(range(len(containers)))

            for rem, indices in remark_index.items():
                # Ambil kontainer sesuai indeks
                sub_containers = [containers[i] for i in indices if i < len(containers)]
                sub_contents = [contents_list[i] for i in indices if i < len(contents_list)]
                sub_meas = [meas_list[i] for i in indices if i < len(meas_list)]
                sub_weight = [weight_list[i] for i in indices if i < len(weight_list)]

                # Ambil prefix dari kontainer pertama (diasumsikan satu grup homogen)
                if sub_containers:
                    prefix = sub_containers[0].split()[0]  # Misal: "PNIU"
                    group_key = (prefix, rem)
                    # Simpan atau overwrite uang tambang per grup
                    uang_tambang_per_grup[group_key] = parse_money(prepaid_raw)

                qty = len(sub_containers)
                if qty == 0:
                    continue
                if qty == 0 or not any([sub_containers, sub_contents, sub_weight]):
                    continue

                key = (no_bl, shipper, consignee, package, rem, kapal, voyage, pel_bongkar, group_key)

                grouped_data[key]["Quantity"] += qty
                grouped_data[key]["Container"].extend(sub_containers)
                grouped_data[key]["Contents"].extend(sub_contents)
                grouped_data[key]["Measurement"].extend(sub_meas)
                grouped_data[key]["Weight"].extend(sub_weight)
                grouped_data[key]["Uang Tambang"] = uang_tambang_per_grup[group_key]

# === Buat DataFrame akhir
final_rows = []
for key, value in grouped_data.items():
    no_bl, shipper, consignee, package, rem, kapal, voyage, pel_bongkar, _ = key
    final_rows.append({
        "No. of B/L": no_bl,
        "Shipper": shipper,
        "Consignee": consignee,
        "Quantity": value["Quantity"],
        "Container": ", ".join(value["Container"]),
        "Package Type": package,
        "Contents": ", ".join(value["Contents"]),
        "Measurement": ", ".join(value["Measurement"]),
        "Weight": ", ".join(value["Weight"]),
        "Uang Tambang": value["Uang Tambang"],
        "Keterangan": rem,
        "Kapal": kapal,
        "Voyage": voyage,
        "Pel. Bongkar": pel_bongkar
    })

df = pd.DataFrame(final_rows)
df.to_csv(output_csv, index=False)
df.to_excel(output_excel, index=False, sheet_name="MANIFEST")

from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# === Buka file Excel yang sudah diexport
wb = load_workbook(output_excel)
ws = wb["MANIFEST"]

# === Definisikan border standar
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# === Terapkan border ke semua cell
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

# === Autofit kolom (menyesuaikan lebar berdasarkan isi)
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            val = str(cell.value)
            if val:
                max_length = max(max_length, len(val))
        except:
            pass
    adjusted_width = max_length + 2  # +2 untuk padding
    ws.column_dimensions[col_letter].width = adjusted_width

# === Simpan kembali
wb.save(output_excel)

print(f"✅ Sukses menyimpan {len(df)} baris ke CSV dan Excel.")

